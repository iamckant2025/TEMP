from flask import Flask, request, send_from_directory
from openpyxl import load_workbook, Workbook
from pathlib import Path

app = Flask(__name__)
BASE_DIR = Path(__file__).resolve().parent
FILE = BASE_DIR / "data.xlsx"

# Create Excel if not exists
if not FILE.exists():
    wb = Workbook()
    ws = wb.active
    ws.title = "FirmData"
    ws["A1"] = "Firm Name"
    ws["A2"] = "GSTIN"
    ws["A3"] = "Address"
    ws["A4"] = "Contact"
    wb.save(FILE)

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/save', methods=['POST'])
def save():
    wb = load_workbook(FILE)
    ws = wb.active
    ws["A1"] = request.form["firm"]
    ws["A2"] = request.form["gstin"]
    ws["A3"] = request.form["address"]
    ws["A4"] = request.form["contact"]
    wb.save(FILE)
    return "<h3>? Data saved successfully!</h3><a href='/'>Go Back</a>"

if __name__ == "__main__":
    app.run(debug=True)
